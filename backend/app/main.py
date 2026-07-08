from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from io import BytesIO
from datetime import datetime
from .pdf_report import crear_pdf
from pathlib import Path

from fastapi.staticfiles import StaticFiles

from fastapi.responses import FileResponse
from .biblioteca import (

    listar_reacciones,

    obtener_reaccion,

    crear_reaccion,

    agregar_variable,

    eliminar_reaccion

)

from typing import List

from fastapi import FastAPI, UploadFile, File, Form

from fastapi.middleware.cors import CORSMiddleware

from .models import DatosEntrada

from .lector import cargar_datos

from .calculos import calcular_parametros

app = FastAPI(
    title="ControladorWeb",
    version="1.0"
)

BASE_DIR = Path(__file__).resolve().parent.parent

FRONTEND = BASE_DIR / "frontend"

app.mount(

    "/css",

    StaticFiles(

        directory=FRONTEND / "css"

    ),

    name="css"

)

app.mount(

    "/js",

    StaticFiles(

        directory=FRONTEND / "js"

    ),

    name="js"

)

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"]
)


@app.get("/")
def inicio():

    return FileResponse(

        FRONTEND / "index.html"

    )


@app.post("/calcular")
def calcular(datos: DatosEntrada):

    data, metadata, archivo = cargar_datos(

        datos.reaccion,

        datos.variable,

        datos.salto


    )

    tiempo_salto = metadata["tiempo_salto"]

    m = metadata["variables"][datos.variable]["m"]

    Kc, KPI, Tint, reglas = calcular_parametros(

        data,

        tiempo_salto,

        m,

        datos.salto


    )

    return {

        "archivo": archivo,

        "tiempo": data[:, 0].tolist(),

        "concentracion": data[:, 1].tolist(),

        "temperaturas": data[:, 2:].tolist(),

        "Kc": Kc,

        "KPI": KPI,

        "Tint": Tint,

        "Reglas": reglas

    }

@app.post("/descargar_excel")
def descargar_excel(datos: DatosEntrada):

    data, metadata, archivo = cargar_datos(
        datos.reaccion,
        datos.variable,
        datos.salto
    )

    tiempo_salto = metadata["tiempo_salto"]

    m = metadata["variables"][datos.variable]["m"]

    Kc, KPI, Tint, reglas = calcular_parametros(

        data,

        tiempo_salto,

        m,

        datos.salto

    )

    # ==========================================
    # Nombres completos
    # ==========================================

    nombre_reaccion = metadata["nombre"]

    reaccion_archivo = datos.reaccion

    nombre_variable = datos.variable

    # ==========================================
    # Crear Excel
    # ==========================================

    wb = Workbook()
    ws = wb.active
    ws.title = "Controlador"

    ws["A1"] = "Reacción"
    ws["B1"] = nombre_reaccion

    ws["A2"] = "Variable"
    ws["B2"] = nombre_variable

    ws["A3"] = "Salto (%)"
    ws["B3"] = datos.salto

    ws["A4"] = "Tiempo del salto"
    ws["B4"] = tiempo_salto

    ws["A5"] = "m"
    ws["B5"] = m

    ws["A6"] = "Fecha"
    ws["B6"] = datetime.now().strftime("%d/%m/%Y %H:%M")

    ws.append([])

    ws.append([
        "Temperatura",
        "Kc",
        "KPI",
        "Tint",
        "Regla"
    ])

    for i in range(11):

        ws.append([
            f"T{i+1}",
            Kc[i],
            KPI[i],
            Tint[i],
            reglas[i]
        ])

    archivo_excel = BytesIO()

    wb.save(archivo_excel)

    archivo_excel.seek(0)

    nombre = (
        f"{reaccion_archivo}_"
        f"{nombre_variable}_"
        f"{datos.salto}.xlsx"
    )

    return StreamingResponse(

        archivo_excel,

        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",

        headers={
            "Content-Disposition":
            f'attachment; filename="{nombre}"'
        }

    )

# =========================================
# GENERAR REPORTE
# =========================================

@app.post("/descargar_informe")
def descargar_informe(datos: dict):

    data, metadata, archivo = cargar_datos(
        datos["reaccion"],
        datos["variable"],
        datos["salto"]
    )

    tiempo_salto = metadata["tiempo_salto"]

    m = metadata["variables"][datos["variable"]]["m"]

    Kc, KPI, Tint, reglas = calcular_parametros(

        data,

        tiempo_salto,

        m,

        datos["salto"]

    )
    pdf = crear_pdf(
        datos,
        metadata,
        Kc,
        KPI,
        Tint,
        reglas
    )

    return StreamingResponse(

        pdf,

        media_type="application/pdf",

        headers={
            "Content-Disposition":
            'attachment; filename="Reporte_Controlador.pdf"'
        }

    )

# =========================================
# BIBLIOTECA DE REACCIONES
# =========================================

@app.get("/reacciones")
def reacciones():

    return listar_reacciones()


@app.get("/reaccion/{nombre}")
def reaccion(nombre: str):

    return obtener_reaccion(nombre)


@app.post("/reaccion")
async def nueva_reaccion(

    nombre: str = Form(...),

    tiempo_salto: float = Form(...),

    m: float = Form(...),

    variable: str = Form(...),

    archivos: List[UploadFile] = File(...)

):

    return await crear_reaccion(

        nombre,

        tiempo_salto,

        m,

        variable,

        archivos

    )

# =========================================
# AGREGAR VARIABLE
# =========================================

@app.post("/reaccion/agregar-variable")
async def nueva_variable(

    nombre: str = Form(...),

    variable: str = Form(...),

    m: float = Form(...),

    archivos: List[UploadFile] = File(...)

):

    return await agregar_variable(

        nombre,

        variable,

        m,

        archivos

    )


# =========================================
# ELIMINAR REACCIÓN
# =========================================

@app.delete("/reaccion/{nombre}")
def borrar_reaccion(

    nombre: str

):

    return eliminar_reaccion(

        nombre

    )
